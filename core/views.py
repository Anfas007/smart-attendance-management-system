from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.core.mail import send_mail
from django.conf import settings
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse # Added JsonResponse
from django.urls import reverse
from django.utils import timezone
from django.contrib import messages # Import messages framework
from django.db.models import Q, Case, When, IntegerField # Add Case and When for ordering
from django.db.models.functions import Substr, Cast, Right
# Make sure all necessary models are imported
from .models import (
    User, AttendanceRecord, Camera, Department, Course, 
    Session, Semester, LeaveRequest, AttendanceSettings
)
from datetime import datetime
import os # For potential file handling if needed later
# --- Imports for Face Recognition ---
import face_recognition
import cv2 # OpenCV for image processing
import numpy as np
import base64
import json
from django.views.decorators.csrf import csrf_exempt # Temporarily for testing API, consider proper CSRF later
from django.views.decorators.http import require_POST
# --- End Face Recognition Imports ---


# -----------------------
# Helper functions
# -----------------------

def is_admin(user):
    """Checks if the user is an authenticated admin."""
    return user.is_authenticated and user.is_admin

def is_student(user):
    """Checks if the user is an authenticated student."""
    return user.is_authenticated and user.is_student

def get_student_form_context(request):
    """Helper to fetch all foreign key options for student forms."""
    # Ensure foreign keys are ordered and available
    return {
        "departments": Department.objects.all().order_by('name'),
        "courses": Course.objects.select_related('department').all().order_by('name'), # Fetch course and dept name together
        "sessions": Session.objects.all().order_by('-year'),
        "semesters": Semester.objects.all().order_by('name'),
    }

# -----------------------
# General views
# -----------------------

def home(request):
    """Renders the homepage."""
    return render(request, "core/home.html")

def user_login(request):
    """Handles user login for both admins and students."""
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)
            messages.success(request, f"Welcome, {user.username}!")
            if user.is_admin:
                return redirect("admin_dashboard")
            elif user.is_student:
                # Add check for authorization if needed
                # if not user.authorized:
                #     logout(request)
                #     messages.warning(request, "Your account is not yet authorized. Please contact an administrator.")
                #     return redirect('login')
                return redirect("student_dashboard")
        else:
            messages.error(request, "Invalid username or password.")
            # No need to pass error context when using messages framework
            return render(request, "core/login.html")

    # For GET request
    return render(request, "core/login.html")

def logout_view(request):
    """Logs out the user."""
    # Clear attendance session if active
    request.session.pop('attendance_session_active', None)
    logout(request)
    messages.info(request, "You have been logged out successfully.")
    return redirect("home")

# -----------------------
# Admin Dashboard & Statistics
# -----------------------

@login_required
@user_passes_test(is_admin)
def admin_dashboard(request):
    """Admin dashboard with key statistics."""
    today = timezone.localdate()
    today_records = AttendanceRecord.objects.filter(date=today)

    # Get leave statistics
    leaves_today = LeaveRequest.objects.filter(
        start_date__lte=today,
        end_date__gte=today,
        status='approved'
    )
    pending_leaves = LeaveRequest.objects.filter(status="pending")

    # Get camera statistics
    all_cameras = Camera.objects.all()
    active_cameras = all_cameras.filter(status='active')

    # Get weekly attendance data (last 7 days)
    weekly_data = []
    for i in range(6, -1, -1):  # Last 7 days including today
        date = today - timezone.timedelta(days=i)
        records = AttendanceRecord.objects.filter(date=date)
        present_count = records.filter(status='present').count()
        absent_count = records.filter(status='absent').count()
        late_count = records.filter(status='late').count()
        weekly_data.append({
            'date': date.strftime('%a'),  # Short day name
            'present': present_count,
            'absent': absent_count,
            'late': late_count,
            'total': present_count + absent_count + late_count
        })

    # Calculate attendance rate
    total_students = User.objects.filter(is_student=True).count()
    present_today = today_records.filter(status="present").count()
    attendance_rate = (present_today / total_students * 100) if total_students > 0 else 0

    context = {
        "total_students": total_students,
        "total_courses": Course.objects.count(),
        "total_cameras": all_cameras.count(),
        "active_cameras": active_cameras.count(),
        "cameras": all_cameras.order_by('-is_default', '-status', 'name')[:6],  # Show top 6 cameras

        # Attendance stats for today
        "total_check_ins": today_records.filter(check_in_time__isnull=False).count(),
        "present": present_today,
        "absent": today_records.filter(status="absent").count(),
        "late": today_records.filter(status="late").count(),
        "attendance_rate": round(attendance_rate, 1),

        # Leave stats
        "students_on_leave": leaves_today.count(),
        "pending_leaves": pending_leaves.count(),
        "recent_leaves": pending_leaves.select_related('student').order_by('-created_at')[:5],

        # Weekly data for charts
        "weekly_data": weekly_data,

        # Include today's date for display
        "today": today
    }
    return render(request, "core/admin_dashboard.html", context)

# -----------------------
# Admin: Student Management (CRUD)
# -----------------------

@login_required
@user_passes_test(is_admin)
def register_student(request):
    """Admin registers a new student with full details."""
    form_context = get_student_form_context(request)

    if request.method == "POST":
        # Get form fields
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        name = request.POST.get("name")
        roll_no = request.POST.get("roll_no")
        contact = request.POST.get("contact")
        department_id = request.POST.get("department")
        course_id = request.POST.get("course")
        session_id = request.POST.get("session")
        semester_id = request.POST.get("semester")
        profile_image = request.FILES.get("profile_image")

        is_active = request.POST.get('is_active') == 'on' if request.POST.get('is_active') is not None else True  # Checkbox value, default to True for new students

        # Validation
        required_fields = [username, password, name, roll_no, department_id, course_id, session_id, semester_id]
        form_context['posted_data'] = request.POST # Repopulate form on error

        # Check if image is provided (either uploaded file or webcam capture)
        profile_image_data = request.POST.get('profile_image_data')
        has_image = profile_image or (profile_image_data and profile_image_data.startswith('data:image'))

        if not has_image:
            messages.error(request, "Profile image is required. Please upload an image or capture one using the webcam.")
            return render(request, "core/add_edit_student.html", form_context)

        if not all(required_fields):
            messages.error(request, "Please fill out all required fields.")
            return render(request, "core/add_edit_student.html", form_context)

        if User.objects.filter(username=username).exists():
            messages.error(request, "A user with this username already exists.")
            return render(request, "core/add_edit_student.html", form_context)

        if User.objects.filter(roll_no=roll_no).exists():
            messages.error(request, "A student with this Roll Number already exists.")
            return render(request, "core/add_edit_student.html", form_context)


        # Create user and save details
        try:
            department = get_object_or_404(Department, pk=department_id)
            course = get_object_or_404(Course, pk=course_id)
            session = get_object_or_404(Session, pk=session_id)
            semester = get_object_or_404(Semester, pk=semester_id)

            user = User.objects.create_user(username=username, email=email, password=password)
            user.name = name
            user.roll_no = roll_no
            user.contact = contact
            user.department = department
            user.course = course
            user.session = session
            user.semester = semester
            user.is_student = True
            user.is_admin = False
            user.is_active = is_active # Set active status

            # Handle either uploaded file or captured image data
            profile_image_data = request.POST.get('profile_image_data')

            if profile_image_data and profile_image_data.startswith('data:image'):
                # Convert base64 image to file
                format, imgstr = profile_image_data.split(';base64,')
                ext = format.split('/')[-1]

                # Create a ContentFile from the decoded base64 data
                from django.core.files.base import ContentFile
                import base64

                decoded_image = base64.b64decode(imgstr)
                from django.utils.text import slugify
                filename = f"{slugify(user.username)}_profile.{ext}"
                user.profile_image.save(filename, ContentFile(decoded_image), save=True)
            elif profile_image:
                user.profile_image = profile_image

            # Check for duplicate faces before saving
            try:
                import face_recognition
                import numpy as np
                from PIL import Image, ImageOps

                # Load and process the image for face detection
                pil_image = Image.open(user.profile_image.path)
                pil_image = ImageOps.exif_transpose(pil_image)

                if pil_image.mode != 'RGB':
                    pil_image = pil_image.convert("RGB")

                # Convert to numpy array for face_recognition
                image_array = np.array(pil_image)

                # --- Image Preprocessing for Better Face Recognition ---
                # Convert to grayscale for preprocessing
                gray = cv2.cvtColor(image_array, cv2.COLOR_RGB2GRAY)

                # Apply histogram equalization to improve contrast
                gray = cv2.equalizeHist(gray)

                # Apply Gaussian blur to reduce noise
                gray = cv2.GaussianBlur(gray, (3, 3), 0)

                # Convert back to RGB
                image_array = cv2.cvtColor(gray, cv2.COLOR_GRAY2RGB)

                # Detect faces
                face_locations = face_recognition.face_locations(image_array, model="hog")
                if not face_locations:
                    # Try CNN model if HOG fails
                    face_locations = face_recognition.face_locations(image_array, model="cnn")

                if face_locations:
                    # Generate encoding for the new image
                    new_encodings = face_recognition.face_encodings(image_array, face_locations)
                    if new_encodings:
                        new_encoding = new_encodings[0]

                        # Compare with existing student encodings
                        existing_students = User.objects.filter(
                            is_student=True,
                            face_encoding__isnull=False
                        ).exclude(face_encoding='')

                        duplicate_found = False
                        for existing_student in existing_students:
                            existing_encoding = existing_student.get_encoding()
                            if existing_encoding is not None:
                                # Compare faces with tolerance
                                matches = face_recognition.compare_faces([existing_encoding], new_encoding, tolerance=0.4)
                                if matches[0]:
                                    duplicate_found = True
                                    messages.error(request, f"This face image is already registered to student '{existing_student.name}' (Roll No: {existing_student.roll_no}). Please use a different image.")
                                    # Delete the created user since registration failed
                                    user.delete()
                                    return render(request, "core/add_edit_student.html", form_context)

                        if not duplicate_found:
                            # Save the encoding for the new student
                            user.set_encoding(new_encoding)
                    else:
                        messages.error(request, "Could not generate face encoding from the uploaded image. Please ensure the image shows a clear face.")
                        user.delete()
                        return render(request, "core/add_edit_student.html", form_context)
                else:
                    messages.error(request, "No face detected in the uploaded image. Please upload a clear photo of your face.")
                    user.delete()
                    return render(request, "core/add_edit_student.html", form_context)

            except Exception as e:
                messages.error(request, f"Error processing image: {str(e)}. Please try again with a different image.")
                user.delete()
                return render(request, "core/add_edit_student.html", form_context)

            user.save()

            messages.info(request, f"Student '{user.name}' registered. Run encoding command to enable face recognition.")

            messages.success(request, f"Student '{user.name}' registered successfully.")
            return redirect("manage_students") # Redirect to student list after success

        except (Department.DoesNotExist, Course.DoesNotExist, Session.DoesNotExist, Semester.DoesNotExist):
             messages.error(request, "Invalid Department, Course, Session, or Semester selected.")
             return render(request, "core/add_edit_student.html", form_context)
        except Exception as e:
             messages.error(request, f"An unexpected error occurred: {e}")
             return render(request, "core/add_edit_student.html", form_context)


    # For GET request, render the empty form with choices
    return render(request, "core/add_edit_student.html", form_context)


@login_required
@user_passes_test(is_admin)
def manage_students_view(request):
    """Displays a list of all registered students with filtering options."""
    students = User.objects.filter(is_student=True).select_related(
        'department', 'course', 'session', 'semester'
    )

    # Get filter parameters
    search_query = request.GET.get('search', '')
    department_id = request.GET.get('department', '')
    course_id = request.GET.get('course', '')
    session_id = request.GET.get('session', '')
    semester_id = request.GET.get('semester', '')
    is_active = request.GET.get('is_active', '')

    # Convert IDs to integers if provided
    try:
        department_id = int(department_id) if department_id else None
    except ValueError:
        department_id = None
    
    try:
        course_id = int(course_id) if course_id else None
    except ValueError:
        course_id = None
        
    try:
        session_id = int(session_id) if session_id else None
    except ValueError:
        session_id = None
        
    try:
        semester_id = int(semester_id) if semester_id else None
    except ValueError:
        semester_id = None

    # Apply filters
    if search_query:
        students = students.filter(
            models.Q(name__icontains=search_query) |
            models.Q(roll_no__icontains=search_query) |
            models.Q(email__icontains=search_query)
        )

    if department_id:
        students = students.filter(department_id=department_id)

    if course_id:
        students = students.filter(course_id=course_id)

    if session_id:
        students = students.filter(session_id=session_id)

    if semester_id:
        students = students.filter(semester_id=semester_id)

    if is_active in ['true', 'false']:
        students = students.filter(is_active=is_active == 'true')

    # Order by the numeric part of roll number (last 4 digits)
    students = students.annotate(
        roll_number_numeric=Cast(Right('roll_no', 4), IntegerField())
    ).order_by('roll_number_numeric')

    # Get all options for dropdown filters
    departments = Department.objects.filter(is_active=True)
    courses = Course.objects.filter(is_active=True)
    sessions = Session.objects.filter(is_active=True)
    semesters = Semester.objects.filter(is_active=True)

    context = {
        "students": students,
        "filters": {
            "search": search_query,
            "department": department_id,
            "course": course_id,
            "session": session_id,
            "semester": semester_id,
            "is_active": is_active
        },
        "filter_options": {
            "departments": departments,
            "courses": courses,
            "sessions": sessions,
            "semesters": semesters
        }
    }
    
    return render(request, "core/manage_students.html", context)

@login_required
@user_passes_test(is_admin)
def edit_student_view(request, user_id):
    """Handles editing an existing student."""
    student = get_object_or_404(User, pk=user_id, is_student=True)
    form_context = get_student_form_context(request)
    form_context['student'] = student
    
    if request.method == 'POST':
        # Get all updated fields
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        name = request.POST.get("name")
        roll_no = request.POST.get("roll_no")
        
        is_active = request.POST.get('is_active') == 'on' # Handle is_active checkbox

        # Check if Roll No changed and conflicts with another user
        if roll_no != student.roll_no and User.objects.filter(roll_no=roll_no).exists():
             messages.error(request, "Another student with this Roll Number already exists.")
             return render(request, "core/add_edit_student.html", form_context)
        
        # Update student object
        student.name = name
        student.email = email
        student.roll_no = roll_no
        student.contact = request.POST.get("contact")
        student.is_active = is_active # Update active status

        # Handle either uploaded file or captured image data
        profile_image_data = request.POST.get('profile_image_data')
        profile_image = request.FILES.get("profile_image")
        
        if profile_image_data and profile_image_data.startswith('data:image'):
            # Convert base64 image to file
            format, imgstr = profile_image_data.split(';base64,')
            ext = format.split('/')[-1]
            
            # Create a ContentFile from the decoded base64 data
            from django.core.files.base import ContentFile
            import base64
            
            decoded_image = base64.b64decode(imgstr)
            from django.utils.text import slugify
            filename = f"{slugify(student.username)}_profile.{ext}"
            student.profile_image.save(filename, ContentFile(decoded_image), save=True)
        elif profile_image:
            student.profile_image = profile_image
        
        # Handle password change only if provided
        if password:
             student.set_password(password)

        try:
             # Handle Foreign Keys
             student.department = get_object_or_404(Department, pk=request.POST.get("department"))
             student.course = get_object_or_404(Course, pk=request.POST.get("course"))
             student.session = get_object_or_404(Session, pk=request.POST.get("session"))
             student.semester = get_object_or_404(Semester, pk=request.POST.get("semester"))
             
             student.save()
             messages.success(request, f"Student '{student.name}' updated successfully. Remember to re-run encoding command if image changed.")
             return redirect('manage_students')
        except Exception as e:
            messages.error(request, f"An error occurred during update: {e}")
            return render(request, "core/add_edit_student.html", form_context)
            
    # For GET request
    return render(request, "core/add_edit_student.html", form_context)


@login_required
@user_passes_test(is_admin)
def authorize_student(request, user_id):
    """Authorizes a student account."""
    student = get_object_or_404(User, pk=user_id, is_student=True)
    if not student.authorized: # Only authorize if not already authorized
        student.authorized = True
        student.save()
        messages.success(request, f"Student '{student.name}' authorized successfully.")
    else:
        messages.info(request, f"Student '{student.name}' is already authorized.")
    return redirect("manage_students")


@login_required
@user_passes_test(is_admin)
def delete_student(request, user_id):
     """Deletes a student account (Requires POST)."""
     if request.method == 'POST':
         student = get_object_or_404(User, pk=user_id, is_student=True)
         try:
             student_name = student.name
             student.delete()
             messages.success(request, f"Student '{student_name}' deleted successfully.")
         except Exception as e:
             messages.error(request, f"Error deleting student: {e}")
     else:
         messages.error(request, "Invalid request method for deletion.")
     return redirect("manage_students")

# -----------------------
# Admin: Course Management
# -----------------------

@login_required
@user_passes_test(is_admin)
def manage_courses_view(request):
    """Displays a list of courses."""
    courses = Course.objects.select_related('department').all().order_by('department__name', 'name')
    context = {"courses": courses}
    return render(request, "core/manage_courses.html", context)

@login_required
@user_passes_test(is_admin)
def add_course_view(request):
    """Handles adding a new course."""
    departments = Department.objects.all().order_by('name') # Fetch departments for the form
    if request.method == 'POST':
        name = request.POST.get('course_name')
        description = request.POST.get('description', '') # Default to empty string
        department_id = request.POST.get('department_id')

        # Validation
        if not name or not department_id:
            messages.error(request, "Course Name and Department are required.")
            return render(request, "core/add_edit_course.html", {
                "departments": departments,
                "course_name": name,
                "description": description,
                "selected_department": int(department_id) if department_id else None
            })

        try:
            department = get_object_or_404(Department, pk=department_id)
            if Course.objects.filter(name=name, department=department).exists():
                messages.error(request, f"Course '{name}' already exists in the {department.name} department.")
                return render(request, "core/add_edit_course.html", {
                    "departments": departments,
                    "course_name": name,
                    "description": description,
                    "selected_department": int(department_id)
                })
            else:
                Course.objects.create(name=name, department=department, description=description)
                messages.success(request, f"Course '{name}' added successfully.")
                return redirect('manage_courses')
        except Department.DoesNotExist:
            messages.error(request, "Selected department not found.")
            return render(request, "core/add_edit_course.html", {
                "departments": departments,
                "course_name": name,
                "description": description,
                "selected_department": int(department_id) if department_id else None
            })

    # For GET request, display the blank form
    return render(request, "core/add_edit_course.html", {"departments": departments})

@login_required
@user_passes_test(is_admin)
def edit_course_view(request, course_id):
    """Handles editing an existing course."""
    course = get_object_or_404(Course, pk=course_id)
    departments = Department.objects.all().order_by('name')
    if request.method == 'POST':
        name = request.POST.get('course_name')
        description = request.POST.get('description', '')
        department_id = request.POST.get('department_id')

        # Validation
        if not name or not department_id:
            messages.error(request, "Course Name and Department are required.")
            return render(request, "core/add_edit_course.html", {
                "course": course,
                "departments": departments,
                "course_name": name, # Pass back potentially invalid data
                "description": description,
                "selected_department": int(department_id) if department_id else None
            })

        try:
            department = get_object_or_404(Department, pk=department_id)
            # Check if name changed and conflicts with another course in the same department
            if course.name != name and Course.objects.filter(name=name, department=department).exclude(pk=course_id).exists():
                messages.error(request, f"Another course named '{name}' already exists in the {department.name} department.")
                return render(request, "core/add_edit_course.html", {
                    "course": course,
                    "departments": departments,
                    "course_name": name,
                    "description": description,
                    "selected_department": int(department_id)
                })
            else:
                course.name = name
                course.department = department
                course.description = description
                course.save()
                messages.success(request, f"Course '{course.name}' updated successfully.")
                return redirect('manage_courses')
        except Department.DoesNotExist:
             messages.error(request, "Selected department not found.")
             return render(request, "core/add_edit_course.html", {
                "course": course,
                "departments": departments,
                "course_name": name,
                "description": description,
                "selected_department": int(department_id) if department_id else None
            })


    # For GET request, display the form pre-filled with course data
    return render(request, "core/add_edit_course.html", {
        "course": course,
        "departments": departments,
        # Pass initial values for the form fields
        "course_name": course.name,
        "description": course.description,
        "selected_department": course.department.id if course.department else None
    })

@login_required
@user_passes_test(is_admin)
def delete_course_view(request, course_id):
    """Handles deleting a course (requires POST)."""
    if request.method == 'POST':
        course = get_object_or_404(Course, pk=course_id)
        try:
            course_name = course.name
            course.delete()
            messages.success(request, f"Course '{course_name}' deleted successfully.")
        except Exception as e:
            # Handle potential deletion issues (e.g., related students) if needed
            messages.error(request, f"Error deleting course '{course.name}': {e}. Check if students are enrolled.")
        return redirect('manage_courses')
    else:
        # Redirect if accessed via GET
        messages.error(request, "Invalid request method for deletion.")
        return redirect('manage_courses')

# -----------------------
# Admin: Department Management
# -----------------------

@login_required
@user_passes_test(is_admin)
def manage_departments_view(request):
    """Displays list of departments."""
    departments = Department.objects.all().order_by('name')
    return render(request, "core/manage_departments.html", {"departments": departments})

@login_required
@user_passes_test(is_admin)
def add_department_view(request):
    """Handles adding a new department."""
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            if Department.objects.filter(name=name).exists():
                messages.error(request, f"Department '{name}' already exists.")
                return render(request, 'core/add_edit_department.html', {'posted_data': request.POST})
            else:
                Department.objects.create(name=name)
                messages.success(request, f"Department '{name}' added successfully.")
                return redirect('manage_departments')
        else:
            messages.error(request, "Department name cannot be empty.")
            return render(request, 'core/add_edit_department.html', {'posted_data': request.POST})
    return render(request, 'core/add_edit_department.html')

@login_required
@user_passes_test(is_admin)
def edit_department_view(request, department_id):
    """Handles editing an existing department."""
    department = get_object_or_404(Department, pk=department_id)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            # Check if name changed and conflicts with another department
            if department.name != name and Department.objects.filter(name=name).exclude(pk=department_id).exists():
                messages.error(request, f"Another department named '{name}' already exists.")
                return render(request, 'core/add_edit_department.html', {
                    'department': department,
                    'posted_data': request.POST # Pass back the problematic data
                })
            else:
                department.name = name
                department.save()
                messages.success(request, f"Department '{department.name}' updated successfully.")
                return redirect('manage_departments')
        else:
            messages.error(request, "Department name cannot be empty.")
            return render(request, 'core/add_edit_department.html', {
                'department': department,
                'posted_data': request.POST
            })
    
    # For GET request, display the form pre-filled with department data
    return render(request, 'core/add_edit_department.html', {'department': department})

@login_required
@user_passes_test(is_admin)
def delete_department_view(request, department_id):
    """Handles deleting a department (requires POST)."""
    if request.method == 'POST':
        department = get_object_or_404(Department, pk=department_id)
        try:
            dept_name = department.name
            department.delete()
            messages.success(request, f"Department '{dept_name}' deleted successfully.")
        except Exception as e:
            # Catch errors if related objects (like Courses or Students) prevent deletion
            messages.error(request, f"Error deleting department '{department.name}': {e}. Make sure no courses or students are assigned.")
        return redirect('manage_departments')
    else:
        messages.error(request, "Invalid request method for deletion.")
        return redirect('manage_departments')


# -----------------------
# Admin: Session Management
# -----------------------

@login_required
@user_passes_test(is_admin)
def manage_sessions_view(request):
    """Displays list of sessions."""
    sessions = Session.objects.all().order_by('-year') # Order newest first
    return render(request, "core/manage_sessions.html", {"sessions": sessions})

@login_required
@user_passes_test(is_admin)
def add_session_view(request):
    """Handles adding a new session."""
    if request.method == 'POST':
        year = request.POST.get('year')
        if year:
            if Session.objects.filter(year=year).exists():
                messages.error(request, f"Session '{year}' already exists.")
                return render(request, 'core/add_edit_session.html', {'posted_data': request.POST})
            else:
                Session.objects.create(year=year)
                messages.success(request, f"Session '{year}' added successfully.")
                return redirect('manage_sessions')
        else:
            messages.error(request, "Session year cannot be empty.")
            return render(request, 'core/add_edit_session.html', {'posted_data': request.POST})

    return render(request, 'core/add_edit_session.html')

@login_required
@user_passes_test(is_admin)
def edit_session_view(request, session_id):
    """Handles editing an existing session."""
    session = get_object_or_404(Session, pk=session_id)
    if request.method == 'POST':
        year = request.POST.get('year', '').strip()
        if year:
            if session.year != year and Session.objects.filter(year=year).exclude(pk=session_id).exists():
                messages.error(request, f"Another session '{year}' already exists.")
                return render(request, 'core/add_edit_session.html', {
                    'session': session,
                    'posted_data': request.POST
                })
            else:
                session.year = year
                session.save()
                messages.success(request, f"Session '{session.year}' updated successfully.")
                return redirect('manage_sessions')
        else:
            messages.error(request, "Session year cannot be empty.")
            return render(request, 'core/add_edit_session.html', {
                'session': session,
                'posted_data': request.POST
            })
    
    # For GET request
    return render(request, 'core/add_edit_session.html', {'session': session})

@login_required
@user_passes_test(is_admin)
def delete_session_view(request, session_id):
    """Handles deleting a session (requires POST)."""
    if request.method == 'POST':
        session = get_object_or_404(Session, pk=session_id)
        try:
            session_year = session.year
            session.delete()
            messages.success(request, f"Session '{session_year}' deleted successfully.")
        except Exception as e:
            messages.error(request, f"Error deleting session '{session.year}': {e}. Check if students are assigned.")
        return redirect('manage_sessions')
    else:
        messages.error(request, "Invalid request method for deletion.")
        return redirect('manage_sessions')


# -----------------------
# Admin: Semester Management
# -----------------------

@login_required
@user_passes_test(is_admin)
def manage_semesters_view(request):
    """Displays list of semesters."""
    semesters = Semester.objects.all().order_by('name')
    return render(request, "core/manage_semesters.html", {"semesters": semesters})

@login_required
@user_passes_test(is_admin)
def add_semester_view(request):
    """Handles adding a new semester."""
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            if Semester.objects.filter(name=name).exists():
                messages.error(request, f"Semester '{name}' already exists.")
                return render(request, 'core/add_edit_semester.html', {'posted_data': request.POST})
            else:
                Semester.objects.create(name=name)
                messages.success(request, f"Semester '{name}' added successfully.")
                return redirect('manage_semesters')
        else:
            messages.error(request, "Semester name cannot be empty.")
            return render(request, 'core/add_edit_semester.html', {'posted_data': request.POST})

    return render(request, 'core/add_edit_semester.html')

@login_required
@user_passes_test(is_admin)
def edit_semester_view(request, semester_id):
    """Handles editing an existing semester."""
    semester = get_object_or_404(Semester, pk=semester_id)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            if semester.name != name and Semester.objects.filter(name=name).exclude(pk=semester_id).exists():
                messages.error(request, f"Another semester named '{name}' already exists.")
                return render(request, 'core/add_edit_semester.html', {
                    'semester': semester,
                    'posted_data': request.POST
                })
            else:
                semester.name = name
                semester.save()
                messages.success(request, f"Semester '{semester.name}' updated successfully.")
                return redirect('manage_semesters')
        else:
            messages.error(request, "Semester name cannot be empty.")
            return render(request, 'core/add_edit_semester.html', {
                'semester': semester,
                'posted_data': request.POST
            })
    
    # For GET request
    return render(request, 'core/add_edit_semester.html', {'semester': semester})

@login_required
@user_passes_test(is_admin)
def delete_semester_view(request, semester_id):
    """Handles deleting a semester (requires POST)."""
    if request.method == 'POST':
        semester = get_object_or_404(Semester, pk=semester_id)
        try:
            semester_name = semester.name
            semester.delete()
            messages.success(request, f"Semester '{semester_name}' deleted successfully.")
        except Exception as e:
            messages.error(request, f"Error deleting semester '{semester.name}': {e}. Check if students are assigned.")
        return redirect('manage_semesters')
    else:
        messages.error(request, "Invalid request method for deletion.")
        return redirect('manage_semesters')


# -----------------------
# Admin: Leave Management
# -----------------------

@login_required
@user_passes_test(is_admin)
def manage_leave_view(request):
    """Displays leave requests with enhanced filtering and course integration."""
    from django.db.models import Q
    print("\n=== Starting manage_leave_view ===")
    print(f"User: {request.user.username}, Is Admin: {request.user.is_admin}")  # Verify admin access
    try:
        if request.method == 'POST':
            request_id = request.POST.get('request_id')
            action = request.POST.get('action')  # 'approve' or 'reject'

            leave_request = get_object_or_404(LeaveRequest, pk=request_id)

            # Only allow action on pending requests
            if leave_request.status != 'pending':
                messages.error(request, "Can only take action on pending requests.")
                return redirect('manage_leave')

            message_type = 'success' if action == 'approve' else 'warning'
            message = None

            if action == 'approve':
                leave_request.status = 'approved'
                message = f"Leave request for {leave_request.student.name} approved."
                messages.success(request, message)
            elif action == 'reject':
                leave_request.status = 'rejected'
                message = f"Leave request for {leave_request.student.name} rejected."
                messages.warning(request, message)
            else:
                messages.error(request, "Invalid action specified.")
                return redirect('manage_leave')

            leave_request.save()

            # Send email notification to student
            try:
                subject = f"Leave Request {leave_request.status.title()}"
                body = f"""Dear {leave_request.student.name},

Your leave request for {leave_request.start_date} to {leave_request.end_date} has been {leave_request.status}.

Reason provided: {leave_request.reason}

If you have any questions, please contact the administration.

Best regards,
Administration Team"""

                if leave_request.student.email:
                    send_mail(
                        subject,
                        body,
                        settings.DEFAULT_FROM_EMAIL,
                        [leave_request.student.email],
                        fail_silently=True
                    )
            except Exception as e:
                print(f"Failed to send email: {e}")

            return redirect('manage_leave')

        # For GET request, display the list with enhanced filtering
        status_filter = request.GET.get('status', 'all')
        date_filter = request.GET.get('date', 'all')
        student_search = request.GET.get('student_search', '')
        course_filter = request.GET.get('course', '')
        department_filter = request.GET.get('department', '')

        # Start with all leave requests with proper select_related
        try:
            print("\n=== Fetching Leave Requests ===")
            
            # First, check total counts without any filters
            all_leaves = LeaveRequest.objects.all()
            total_count = all_leaves.count()
            print(f"\nDatabase Status:")
            print(f"Total leave requests in system: {total_count}")
            
            if total_count == 0:
                print("\nNO LEAVE REQUESTS FOUND IN DATABASE!")
            else:
                print("\nExisting Leave Requests:")
                for leave in all_leaves:
                    print(f"- ID: {leave.id}, Student: {leave.student.username}, "
                          f"Status: {leave.status}, Created: {leave.created_at}")

            # Now get the filtered queryset with all related data
            leave_requests = LeaveRequest.objects.select_related(
                'student',
                'student__course',
                'student__department'
            ).all().order_by('-created_at')

            print(f"\nFiltered Query Results:")
            print(f"Leave requests after select_related: {leave_requests.count()}")
            
            # Print the SQL query for debugging
            print("\nSQL Query:")
            print(str(leave_requests.query))
            
            # Apply filters
            if status_filter and status_filter != 'all':
                leave_requests = leave_requests.filter(status=status_filter)
                print(f"After status filter ({status_filter}): {leave_requests.count()}")  # Debug print
        except Exception as e:
            print(f"ERROR in queryset: {str(e)}")
            raise

        if student_search:
            leave_requests = leave_requests.filter(
                models.Q(student__name__icontains=student_search) |
                models.Q(student__roll_no__icontains=student_search)
            )
            
        # Add course and department filters
        if course_filter:
            leave_requests = leave_requests.filter(student__course_id=course_filter)
            
        if department_filter:
            leave_requests = leave_requests.filter(student__department_id=department_filter)

        # Apply date filter
        today = timezone.localdate()
        if date_filter == 'upcoming':
            leave_requests = leave_requests.filter(start_date__gte=today)
        elif date_filter == 'past':
            leave_requests = leave_requests.filter(end_date__lt=today)
        elif date_filter == 'current':
            leave_requests = leave_requests.filter(
                start_date__lte=today,
                end_date__gte=today
            )

        # Order by status (pending first) and created date
        leave_requests = leave_requests.order_by(
            Case(
                When(status='pending', then=0),
                When(status='approved', then=1),
                When(status='rejected', then=2),
                default=3,
            ),
            '-created_at'
        )

        # Get summary counts
        summary = {
            'total': leave_requests.count(),
            'pending': leave_requests.filter(status='pending').count(),
            'approved': leave_requests.filter(status='approved').count(),
            'rejected': leave_requests.filter(status='rejected').count(),
            'current': leave_requests.filter(
                status='approved',
                start_date__lte=today,
                end_date__gte=today
            ).count()
        }

        # Get all active courses and departments for filtering
        all_courses = Course.objects.filter(is_active=True).select_related('department')
        all_departments = Department.objects.filter(is_active=True)

        # Enhanced summary with course and department stats
        course_stats = {}
        department_stats = {}
        
        for leave in leave_requests:
            if leave.student.course_id:
                course_stats[leave.student.course_id] = course_stats.get(leave.student.course_id, 0) + 1
            if leave.student.department_id:
                department_stats[leave.student.department_id] = department_stats.get(leave.student.department_id, 0) + 1

        # Debug information about the results
        print(f"Final leave requests count: {leave_requests.count()}")
        print(f"Filters applied - Status: {status_filter}, Date: {date_filter}, Course: {course_filter}, Department: {department_filter}")
        
        # Debug information before creating context
        print("\n=== Context Debug Info ===")
        print(f"Leave Requests Count: {leave_requests.count()}")
        print(f"Summary: {summary}")
        
        context = {
            "leave_requests": leave_requests,
            "summary": summary,
            "debug_info": {
                "total_requests": leave_requests.count(),
                "request_ids": list(leave_requests.values_list('id', flat=True)),
                "has_requests": leave_requests.exists(),
            },
            "filters": {
                "status": status_filter,
                "date": date_filter,
                "student_search": student_search,
                "course": course_filter,
                "department": department_filter
            },
            "filter_options": {
                "courses": all_courses,
                "departments": all_departments,
                "course_stats": course_stats,
                "department_stats": department_stats
            },
            "status_choices": [
                ('all', 'All Requests'),
                ('pending', 'Pending'),
                ('approved', 'Approved'),
                ('rejected', 'Rejected')
            ],
            "date_choices": [
                ('all', 'All Dates'),
                ('current', 'Currently Active'),
                ('upcoming', 'Upcoming'),
                ('past', 'Past')
            ]
        }

        return render(request, "core/manage_leave.html", context)

    except Exception as e:
        print(f"Error in manage_leave_view: {e}")
        messages.error(request, "An error occurred while retrieving leave requests. Please try again.")
        return redirect('admin_dashboard')

# -----------------------
# Admin: Camera Management
# -----------------------

@login_required
@user_passes_test(is_admin)
def manage_cameras_view(request):
    """Displays list of cameras with enhanced management features."""
    cameras = Camera.objects.all().order_by('name')

    # Get filter parameters
    status_filter = request.GET.get('status', 'all')
    type_filter = request.GET.get('type', 'all')

    # Apply filters
    if status_filter and status_filter != 'all':
        cameras = cameras.filter(status=status_filter)
    if type_filter and type_filter != 'all':
        cameras = cameras.filter(camera_type=type_filter)

    context = {
        "cameras": cameras,
        "filters": {
            "status": status_filter,
            "type": type_filter,
        },
        "status_choices": Camera.STATUS_CHOICES,
        "type_choices": Camera.CAMERA_TYPES,
    }

    return render(request, "core/manage_cameras.html", context)

@login_required
@user_passes_test(is_admin)
def add_camera_view(request):
    """Handles adding a new camera with enhanced configuration."""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        camera_type = request.POST.get('camera_type', 'usb')
        location = request.POST.get('location', '').strip()
        ip_address = request.POST.get('ip_address', '').strip() or None
        port = request.POST.get('port') or None
        stream_url = request.POST.get('stream_url', '').strip() or None
        device_index = request.POST.get('device_index', '0')
        status = request.POST.get('status', 'inactive')
        is_default = request.POST.get('is_default') == 'on'
        resolution_width = request.POST.get('resolution_width', '640')
        resolution_height = request.POST.get('resolution_height', '480')
        fps = request.POST.get('fps', '30')

        # Validation
        if not name or not location:
            messages.error(request, "Camera Name and Location are required.")
            return render(request, 'core/add_edit_camera.html', {
                'posted_data': request.POST,
                'form_values': {
                    'name': name,
                    'camera_type': camera_type,
                    'location': location,
                    'ip_address': ip_address,
                    'port': port,
                    'stream_url': stream_url,
                    'device_index': device_index,
                    'status': status,
                    'is_default': is_default,
                    'resolution_width': resolution_width,
                    'resolution_height': resolution_height,
                    'fps': fps,
                }
            })

        if Camera.objects.filter(name=name).exists():
            messages.error(request, f"Camera named '{name}' already exists.")
            return render(request, 'core/add_edit_camera.html', {
                'posted_data': request.POST,
                'form_values': {
                    'name': name,
                    'camera_type': camera_type,
                    'location': location,
                    'ip_address': ip_address,
                    'port': port,
                    'stream_url': stream_url,
                    'device_index': device_index,
                    'status': status,
                    'is_default': is_default,
                    'resolution_width': resolution_width,
                    'resolution_height': resolution_height,
                    'fps': fps,
                }
            })

        # Type-specific validation
        if camera_type == 'ip':
            if not ip_address:
                messages.error(request, "IP Address is required for IP cameras.")
                return render(request, 'core/add_edit_camera.html', {
                    'posted_data': request.POST,
                    'form_values': {
                        'name': name,
                        'camera_type': camera_type,
                        'location': location,
                        'ip_address': ip_address,
                        'port': port,
                        'stream_url': stream_url,
                        'device_index': device_index,
                        'status': status,
                        'is_default': is_default,
                        'resolution_width': resolution_width,
                        'resolution_height': resolution_height,
                        'fps': fps,
                    }
                })
            if not stream_url:
                messages.error(request, "Stream URL is required for IP cameras.")
                return render(request, 'core/add_edit_camera.html', {
                    'posted_data': request.POST,
                    'form_values': {
                        'name': name,
                        'camera_type': camera_type,
                        'location': location,
                        'ip_address': ip_address,
                        'port': port,
                        'stream_url': stream_url,
                        'device_index': device_index,
                        'status': status,
                        'is_default': is_default,
                        'resolution_width': resolution_width,
                        'resolution_height': resolution_height,
                        'fps': fps,
                    }
                })

        try:
            # Convert numeric fields
            if port:
                port = int(port)
            device_index = int(device_index)
            resolution_width = int(resolution_width)
            resolution_height = int(resolution_height)
            fps = int(fps)

            camera = Camera.objects.create(
                name=name,
                camera_type=camera_type,
                location=location,
                ip_address=ip_address,
                port=port,
                stream_url=stream_url,
                device_index=device_index,
                status=status,
                is_default=is_default,
                resolution_width=resolution_width,
                resolution_height=resolution_height,
                fps=fps,
            )

            messages.success(request, f"Camera '{name}' added successfully.")
            return redirect('manage_cameras')

        except ValueError as e:
            messages.error(request, f"Invalid numeric value: {e}")
            return render(request, 'core/add_edit_camera.html', {
                'posted_data': request.POST,
                'form_values': {
                    'name': name,
                    'camera_type': camera_type,
                    'location': location,
                    'ip_address': ip_address,
                    'port': port,
                    'stream_url': stream_url,
                    'device_index': device_index,
                    'status': status,
                    'is_default': is_default,
                    'resolution_width': resolution_width,
                    'resolution_height': resolution_height,
                    'fps': fps,
                }
            })
        except Exception as e:
            messages.error(request, f"Error creating camera: {e}")
            return render(request, 'core/add_edit_camera.html', {
                'posted_data': request.POST,
                'form_values': {
                    'name': name,
                    'camera_type': camera_type,
                    'location': location,
                    'ip_address': ip_address,
                    'port': port,
                    'stream_url': stream_url,
                    'device_index': device_index,
                    'status': status,
                    'is_default': is_default,
                    'resolution_width': resolution_width,
                    'resolution_height': resolution_height,
                    'fps': fps,
                }
            })

    return render(request, 'core/add_edit_camera.html')


@login_required
@user_passes_test(is_admin)
def edit_camera_view(request, camera_id):
    """Handles editing an existing camera with enhanced configuration."""
    camera = get_object_or_404(Camera, pk=camera_id)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        camera_type = request.POST.get('camera_type', camera.camera_type)
        location = request.POST.get('location', '').strip()
        ip_address = request.POST.get('ip_address', '').strip() or None
        port = request.POST.get('port') or None
        stream_url = request.POST.get('stream_url', '').strip() or None
        device_index = request.POST.get('device_index', str(camera.device_index))
        status = request.POST.get('status', camera.status)
        is_default = request.POST.get('is_default') == 'on'
        resolution_width = request.POST.get('resolution_width', str(camera.resolution_width))
        resolution_height = request.POST.get('resolution_height', str(camera.resolution_height))
        fps = request.POST.get('fps', str(camera.fps))

        # Validation
        if not name or not location:
            messages.error(request, "Camera Name and Location are required.")
            return render(request, 'core/add_edit_camera.html', {
                'camera': camera,
                'posted_data': request.POST,
                'form_values': {
                    'name': name,
                    'camera_type': camera_type,
                    'location': location,
                    'ip_address': ip_address,
                    'port': port,
                    'stream_url': stream_url,
                    'device_index': device_index,
                    'status': status,
                    'is_default': is_default,
                    'resolution_width': resolution_width,
                    'resolution_height': resolution_height,
                    'fps': fps,
                }
            })

        # Check name uniqueness (excluding current camera)
        if Camera.objects.filter(name=name).exclude(pk=camera_id).exists():
            messages.error(request, f"Another camera named '{name}' already exists.")
            return render(request, 'core/add_edit_camera.html', {
                'camera': camera,
                'posted_data': request.POST,
                'form_values': {
                    'name': name,
                    'camera_type': camera_type,
                    'location': location,
                    'ip_address': ip_address,
                    'port': port,
                    'stream_url': stream_url,
                    'device_index': device_index,
                    'status': status,
                    'is_default': is_default,
                    'resolution_width': resolution_width,
                    'resolution_height': resolution_height,
                    'fps': fps,
                }
            })

        # Type-specific validation
        if camera_type == 'ip':
            if not ip_address:
                messages.error(request, "IP Address is required for IP cameras.")
                return render(request, 'core/add_edit_camera.html', {
                    'camera': camera,
                    'posted_data': request.POST,
                    'form_values': {
                        'name': name,
                        'camera_type': camera_type,
                        'location': location,
                        'ip_address': ip_address,
                        'port': port,
                        'stream_url': stream_url,
                        'device_index': device_index,
                        'status': status,
                        'is_default': is_default,
                        'resolution_width': resolution_width,
                        'resolution_height': resolution_height,
                        'fps': fps,
                    }
                })
            if not stream_url:
                messages.error(request, "Stream URL is required for IP cameras.")
                return render(request, 'core/add_edit_camera.html', {
                    'camera': camera,
                    'posted_data': request.POST,
                    'form_values': {
                        'name': name,
                        'camera_type': camera_type,
                        'location': location,
                        'ip_address': ip_address,
                        'port': port,
                        'stream_url': stream_url,
                        'device_index': device_index,
                        'status': status,
                        'is_default': is_default,
                        'resolution_width': resolution_width,
                        'resolution_height': resolution_height,
                        'fps': fps,
                    }
                })

        try:
            # Convert numeric fields
            if port:
                port = int(port)
            device_index = int(device_index)
            resolution_width = int(resolution_width)
            resolution_height = int(resolution_height)
            fps = int(fps)

            # Update camera
            camera.name = name
            camera.camera_type = camera_type
            camera.location = location
            camera.ip_address = ip_address
            camera.port = port
            camera.stream_url = stream_url
            camera.device_index = device_index
            camera.status = status
            camera.is_default = is_default
            camera.resolution_width = resolution_width
            camera.resolution_height = resolution_height
            camera.fps = fps
            camera.save()

            messages.success(request, f"Camera '{name}' updated successfully.")
            return redirect('manage_cameras')

        except ValueError as e:
            messages.error(request, f"Invalid numeric value: {e}")
            return render(request, 'core/add_edit_camera.html', {
                'camera': camera,
                'posted_data': request.POST,
                'form_values': {
                    'name': name,
                    'camera_type': camera_type,
                    'location': location,
                    'ip_address': ip_address,
                    'port': port,
                    'stream_url': stream_url,
                    'device_index': device_index,
                    'status': status,
                    'is_default': is_default,
                    'resolution_width': resolution_width,
                    'resolution_height': resolution_height,
                    'fps': fps,
                }
            })
        except Exception as e:
            messages.error(request, f"Error updating camera: {e}")
            return render(request, 'core/add_edit_camera.html', {
                'camera': camera,
                'posted_data': request.POST,
                'form_values': {
                    'name': name,
                    'camera_type': camera_type,
                    'location': location,
                    'ip_address': ip_address,
                    'port': port,
                    'stream_url': stream_url,
                    'device_index': device_index,
                    'status': status,
                    'is_default': is_default,
                    'resolution_width': resolution_width,
                    'resolution_height': resolution_height,
                    'fps': fps,
                }
            })

    return render(request, 'core/add_edit_camera.html', {'camera': camera})


@login_required
@user_passes_test(is_admin)
def toggle_camera_status(request, camera_id):
    """Toggle camera active/inactive status."""
    if request.method == 'POST':
        camera = get_object_or_404(Camera, pk=camera_id)
        if camera.status == 'active':
            camera.status = 'inactive'
            messages.success(request, f"Camera '{camera.name}' deactivated.")
        else:
            camera.status = 'active'
            messages.success(request, f"Camera '{camera.name}' activated.")
        camera.save()
    return redirect('manage_cameras')


@login_required
@user_passes_test(is_admin)
def set_default_camera(request, camera_id):
    """Set a camera as the default camera."""
    if request.method == 'POST':
        camera = get_object_or_404(Camera, pk=camera_id)
        # This will automatically unset other default cameras due to the save() method
        camera.is_default = True
        camera.save()
        messages.success(request, f"Camera '{camera.name}' set as default.")
    return redirect('manage_cameras')


@login_required
@user_passes_test(is_admin)
def test_camera_connection(request, camera_id):
    """Test camera connection and update status."""
    camera = get_object_or_404(Camera, pk=camera_id)

    try:
        # For now, we'll just simulate a connection test
        # In a real implementation, you'd test the actual camera connection
        if camera.camera_type == 'usb':
            # Test USB camera availability
            import cv2
            cap = cv2.VideoCapture(camera.device_index)
            if cap.isOpened():
                camera.status = 'active'
                messages.success(request, f"Camera '{camera.name}' connection test successful.")
                cap.release()
            else:
                camera.status = 'offline'
                messages.error(request, f"Camera '{camera.name}' not accessible.")
        else:
            # For IP cameras, we'd test the stream URL
            # This is a placeholder - real implementation would test the RTSP/HTTP stream
            camera.status = 'active'
            messages.success(request, f"Camera '{camera.name}' connection test simulated.")

        camera.save()

    except Exception as e:
        camera.status = 'offline'
        camera.save()
        messages.error(request, f"Error testing camera '{camera.name}': {e}")

    return redirect('manage_cameras')


@login_required
@user_passes_test(is_admin)
def delete_camera_view(request, camera_id):
    """Handles deleting a camera (requires POST)."""
    if request.method == 'POST':
        camera = get_object_or_404(Camera, pk=camera_id)
        try:
            camera_name = camera.name
            camera.delete()
            messages.success(request, f"Camera '{camera_name}' deleted successfully.")
        except Exception as e:
            messages.error(request, f"Error deleting camera: {e}")
        return redirect('manage_cameras')
    else:
        messages.error(request, "Invalid request method for deletion.")
        return redirect('manage_cameras')


# -----------------------
# Admin: Attendance Management
# -----------------------

@login_required
@user_passes_test(is_admin)
def attendance_details_view(request):
    """Displays attendance records with mandatory department, course, and semester filtering."""
    
    # Check if we should clear the course/semester filter
    if request.GET.get('clear_session_filter') == '1':
        request.session.pop('selected_course_id', None)
        request.session.pop('selected_semester_id', None)
        return redirect('attendance_details')
    
    # Get filter parameters from request
    student_search = request.GET.get('student_search', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    status = request.GET.get('status', '')
    department_id = request.GET.get('department', '')
    course_id = request.GET.get('course', '')
    session_id = request.GET.get('session', '')
    semester_id = request.GET.get('semester', '')

    # Start with base queryset - but only if we have the required filters
    has_required_filters = department_id and course_id and semester_id
    
    if has_required_filters:
        records = AttendanceRecord.objects.all().select_related('student__department', 'student__course', 'student__session', 'student__semester')
        
        # Apply mandatory filters first
        records = records.filter(
            student__department_id=department_id,
            student__course_id=course_id,
            student__semester_id=semester_id
        )
        
        # Apply additional filters
        if student_search:
            records = records.filter(
                models.Q(student__name__icontains=student_search) |
                models.Q(student__roll_no__icontains=student_search)
            )

        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                records = records.filter(date__gte=start_date)
            except ValueError:
                messages.error(request, "Invalid start date format. Please use DD/MM/YYYY.")

        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                records = records.filter(date__lte=end_date)
            except ValueError:
                messages.error(request, "Invalid end date format. Please use DD/MM/YYYY.")

        if status and status in dict(AttendanceRecord.STATUS_CHOICES):
            records = records.filter(status=status)

        if session_id:
            records = records.filter(student__session_id=session_id)

        # Order by date and student roll number (numeric part)
        records = records.annotate(
            roll_number_numeric=Cast(Right('student__roll_no', 4), IntegerField())
        ).order_by('-date', 'roll_number_numeric')
    else:
        # No required filters applied - show empty queryset
        records = AttendanceRecord.objects.none()

    # Get filter options
    departments = Department.objects.filter(is_active=True)
    courses = Course.objects.filter(is_active=True)
    sessions = Session.objects.filter(is_active=True)
    semesters = Semester.objects.filter(is_active=True)

    context = {
        "records": records,
        "has_required_filters": has_required_filters,
        "filters": {
            "student_search": student_search,
            "start_date": start_date,
            "end_date": end_date,
            "status": status,
            "department": department_id,
            "course": course_id,
            "session": session_id,
            "semester": semester_id,
            "status_choices": AttendanceRecord.STATUS_CHOICES
        },
        "departments": departments,
        "courses": courses,
        "sessions": sessions,
        "semesters": semesters,
    }
    
    return render(request, "core/attendance_details.html", context)

# Note: mark_attendance view was removed and merged into face_attendance_view

@login_required
@user_passes_test(is_admin)
def end_attendance_session_view(request):
    """Ends the active attendance session with admin credential verification."""
    if request.method == 'POST':
        # Verify admin password
        password = request.POST.get('password', '')
        if not password:
            messages.error(request, "Please enter your password to end the attendance session.")
            return render(request, 'core/end_session.html', {'attendance_session_active': True})

        # Check if password is correct
        if not request.user.check_password(password):
            messages.error(request, "Incorrect password. Please try again.")
            return render(request, 'core/end_session.html', {'attendance_session_active': True})

        # Password verified, end the session
        request.session.pop('attendance_session_active', None)
        request.session.pop('selected_course_id', None)
        request.session.pop('selected_semester_id', None)
        request.session.modified = True
        messages.success(request, "Attendance session ended successfully.")
        return redirect('admin_dashboard')

    # GET request - show password confirmation form
    return render(request, 'core/end_session.html', {'attendance_session_active': True})

@login_required
@user_passes_test(is_admin)
def notify_absent_or_late(request):
    """Sends email notifications to students marked absent or late today (POST only)."""
    if request.method == 'POST':
        today = timezone.localdate()
        # Fetch students who are marked absent or late today
        records = AttendanceRecord.objects.filter(date=today, status__in=['absent', 'late']).select_related('student')
        count = 0
        failed = 0
        
        if not records.exists():
             messages.info(request, "No students marked absent or late today.")
             return redirect('admin_dashboard')

        for rec in records:
            if rec.student.email:
                subject = f"Attendance Alert: Marked {rec.status.capitalize()} Today ({today})"
                body = f"Hello {rec.student.name},\n\nThis is a notification that your attendance status for {today} has been marked as '{rec.status.capitalize()}'.\n\nPlease contact the administration if you believe this is an error.\n\nThank you."
                try:
                    send_mail(subject, body, settings.DEFAULT_FROM_EMAIL, [rec.student.email], fail_silently=False)
                    count += 1
                except Exception as e:
                    failed += 1
                    print(f"Failed to send email to {rec.student.email}: {e}") # Log error to console/log file

        # Provide feedback based on outcome
        if failed > 0 and count > 0:
            messages.warning(request, f"{count} notification(s) sent successfully. {failed} failed to send.")
        elif failed > 0 and count == 0:
             messages.error(request, f"Failed to send {failed} notification(s). Check email settings and logs.")
        elif count > 0:
            messages.success(request, f"{count} notification(s) sent successfully.")
        # No else needed as the 'no records' case is handled above

        return redirect('admin_dashboard')

    # If accessed via GET, show error and redirect
    messages.error(request, "Invalid request method. Use the button on the dashboard.")
    return redirect('admin_dashboard')

# -----------------------
# Admin Action Views
# -----------------------

@login_required
@user_passes_test(is_admin)
def authorize_student(request, user_id):
    """Authorizes a student account."""
    # This might be better handled via POST from manage_students page for security
    student = get_object_or_404(User, pk=user_id, is_student=True)
    if not student.authorized: # Only authorize if not already authorized
        student.authorized = True
        student.save()
        messages.success(request, f"Student '{student.name}' authorized successfully.")
        # Optionally send notification email
    else:
        messages.info(request, f"Student '{student.name}' is already authorized.")
    return redirect("manage_students")


@login_required
@user_passes_test(is_admin)
def delete_student(request, user_id):
     """Deletes a student account (Requires POST)."""
     # Should ideally be POST request with confirmation
     if request.method == 'POST':
         student = get_object_or_404(User, pk=user_id, is_student=True)
         try:
             student_name = student.name
             student.delete()
             messages.success(request, f"Student '{student_name}' deleted successfully.")
         except Exception as e:
             messages.error(request, f"Error deleting student: {e}")
     else:
         messages.error(request, "Invalid request method for deletion.")
     return redirect("manage_students")

# -----------------------
# Face Recognition Views
# -----------------------

@login_required
@user_passes_test(is_admin)
def mark_absent_students(request=None, specific_date=None):
    """Mark students as absent if they haven't been marked present/late for the day."""
    date_to_mark = specific_date or timezone.localdate()
    
    # Get all active students
    active_students = User.objects.filter(is_student=True, is_active=True)
    
    # Get students who already have attendance records for the day
    marked_students = AttendanceRecord.objects.filter(
        date=date_to_mark
    ).values_list('student_id', flat=True)
    
    # Get students with approved leave requests
    on_leave_students = LeaveRequest.objects.filter(
        start_date__lte=date_to_mark,
        end_date__gte=date_to_mark,
        status='approved'
    ).values_list('student_id', flat=True)
    
    # Find students without attendance records and not on leave
    unmarked_students = active_students.exclude(
        id__in=marked_students
    ).exclude(
        id__in=on_leave_students
    )
    
    # Mark these students as absent
    records_to_create = [
        AttendanceRecord(
            student=student,
            date=date_to_mark,
            status='absent',
            manually_marked=False
        )
        for student in unmarked_students
    ]
    
    if records_to_create:
        AttendanceRecord.objects.bulk_create(records_to_create)
        if request:
            messages.info(request, f"Marked {len(records_to_create)} student(s) as absent.")
    elif request:
        messages.info(request, "No students to mark as absent.")
    
    return len(records_to_create) if records_to_create else 0

@login_required
@user_passes_test(is_admin)
def face_attendance_view(request):
    """Renders the page for face recognition camera feed and manual marking."""
    # Start attendance session - lock admin into this page
    request.session['attendance_session_active'] = True
    request.session.modified = True

    # Get course and semester selection
    selected_course_id = request.GET.get('course_id') or request.session.get('selected_course_id')
    selected_semester_id = request.GET.get('semester_id') or request.session.get('selected_semester_id')

    # Get all available courses and semesters
    courses = Course.objects.filter(is_active=True)
    semesters = Semester.objects.filter(is_active=True)

    # If course and semester are not selected, show selection form
    if not selected_course_id or not selected_semester_id:
        context = {
            'courses': courses,
            'semesters': semesters,
            'attendance_session_active': True,
            'show_course_selection': True,
        }
        return render(request, 'core/face_attendance.html', context)

    # Store selection in session
    request.session['selected_course_id'] = selected_course_id
    request.session['selected_semester_id'] = selected_semester_id
    request.session.modified = True

    # Get selected course and semester objects
    try:
        selected_course = Course.objects.get(id=selected_course_id, is_active=True)
        selected_semester = Semester.objects.get(id=selected_semester_id, is_active=True)
    except (Course.DoesNotExist, Semester.DoesNotExist):
        messages.error(request, "Invalid course or semester selection.")
        return redirect('face_attendance')

    # Filter students by selected course and semester
    students = User.objects.filter(
        is_student=True,
        authorized=True,
        course_id=selected_course_id,
        semester_id=selected_semester_id
    ).select_related('department', 'course', 'session', 'semester').annotate(
        roll_number_numeric=Cast(Right('roll_no', 4), IntegerField())
    ).order_by('roll_number_numeric')

    time_settings = AttendanceSettings.get_instance()

    # Get all active cameras for selection
    cameras = Camera.objects.all().order_by('-is_default', '-status', 'name')

    context = {
        "students": students,
        "settings": time_settings,
        "cameras": cameras,
        "attendance_session_active": True,
        "selected_course": selected_course,
        "selected_semester": selected_semester,
        "courses": courses,
        "semesters": semesters,
        # Initialize form values for manual marking
        "selected_student": None,
        "selected_date": "",
        "selected_status": ""
    }

    # --- Handle Manual Marking POST within this view ---
    if request.method == "POST":
        if 'mark_manual_attendance' in request.POST:
            student_id = request.POST.get("student_id")
            date_str = request.POST.get("date")
            status = request.POST.get("status")
            current_time = timezone.localtime().time()

            context['selected_student'] = int(student_id) if student_id else None
            context['selected_date'] = date_str
            context['selected_status'] = status

            if student_id and date_str and status:
                try:
                    date = datetime.strptime(date_str, "%Y-%m-%d").date()
                    student = get_object_or_404(User, pk=student_id, is_student=True)
                    defaults = {
                        'status': status,
                        'manually_marked': True,
                        'check_in_time': current_time
                    }

                    record, created = AttendanceRecord.objects.update_or_create(
                        student=student,
                        date=date,
                        defaults=defaults
                    )
                    if created:
                        messages.success(request, f"Attendance manually marked for {student.name} on {date} as {status}.")
                    else:
                        messages.success(request, f"Attendance manually updated for {student.name} on {date} to {status}.")
                    return redirect("face_attendance")
                except ValueError:
                    messages.error(request, "Invalid date format. Please use DD/MM/YYYY.")
                except User.DoesNotExist:
                    messages.error(request, "Selected student not found.")
                except Exception as e:
                    messages.error(request, f"An error occurred during manual marking: {e}")
            else:
                messages.error(request, "Please select student, date, and status for manual marking.")
            return render(request, 'core/face_attendance.html', context)

    return render(request, 'core/face_attendance.html', context)


# --- UPDATED: load_known_faces to use stored encodings ---
known_face_encodings_global = []
known_face_ids_global = []

def load_known_faces():
    """
    Loads face encodings and corresponding User IDs from the database.
    This is more efficient than loading images. Run generate_encodings command first.
    """
    global known_face_encodings_global, known_face_ids_global
    
    known_face_encodings_global = []
    known_face_ids_global = []
    
    # Base query for students with encodings
    students_with_encodings = User.objects.filter(
        is_student=True, 
        authorized=True, 
        face_encoding__isnull=False
    ).exclude(face_encoding='')
    
    print(f"[Face Load] Loading encodings for {students_with_encodings.count()} students from DB...") 
    
    for student in students_with_encodings:
        encoding = student.get_encoding() # Use the method from the User model
        if encoding is not None:
            known_face_encodings_global.append(encoding)
            known_face_ids_global.append(student.id)
        else:
            # This case indicates an error during decoding in get_encoding()
             print(f"[Face Load Warning] Could not decode encoding for {student.username}")
            
    print(f"[Face Load] Loaded {len(known_face_encodings_global)} known encodings from DB.")

# --- Load faces on server start ---
# NOTE: Removed the call from here to prevent issues during migrations/checks
# We will load it lazily in the recognize_face_view if it's empty
# load_known_faces() 

# --- VIEW FOR CHECKOUT API ---
@csrf_exempt
@require_POST
@login_required
@user_passes_test(is_admin)
def checkout_student_view(request):
    """Handle student check-out."""
    
    # Get selected course and semester from session
    selected_course_id = request.session.get('selected_course_id')
    selected_semester_id = request.session.get('selected_semester_id')
    
    if not selected_course_id or not selected_semester_id:
        return JsonResponse({
            'status': 'error', 
            'message': 'No course and semester selected for attendance session.'
        }, status=400)
    
    try:
        data = json.loads(request.body)
        image_data_uri = data.get('image_data')

        if not image_data_uri or ',' not in image_data_uri:
            return JsonResponse({'status': 'error', 'message': 'Invalid image data format.'}, status=400)

        # Decode and process the image
        try:
            header, encoded = image_data_uri.split(',', 1)
            image_data = base64.b64decode(encoded)
        except (ValueError, TypeError) as e:
            return JsonResponse({'status': 'error', 'message': f'Error decoding base64 image: {e}'}, status=400)

        # Convert to OpenCV format
        nparr = np.frombuffer(image_data, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        if frame is None:
            return JsonResponse({'status': 'error', 'message': 'Could not decode image data.'}, status=400)

        # Convert BGR to RGB
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        
        # Find faces in the image
        face_locations = face_recognition.face_locations(rgb_frame, model="hog")
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        if not face_encodings:
            return JsonResponse({'status': 'no_face', 'message': 'No face detected.'})

        # Compare with known faces
        if not known_face_encodings_global:
            load_known_faces()

        matches = face_recognition.compare_faces(known_face_encodings_global, face_encodings[0], tolerance=0.45)
        face_distances = face_recognition.face_distance(known_face_encodings_global, face_encodings[0])

        if len(face_distances) > 0:
            best_match_index = np.argmin(face_distances)
            best_distance = face_distances[best_match_index]
            
            # Debug logging for checkout
            print(f"[Checkout Debug] Best match index: {best_match_index}")
            print(f"[Checkout Debug] Best distance: {best_distance:.4f}")
            print(f"[Checkout Debug] Tolerance check: {best_distance <= 0.45}")
            
            if matches[best_match_index]:
                student_id = known_face_ids_global[best_match_index]
                try:
                    student = User.objects.get(id=student_id)
                    
                    # VALIDATE: Check if student belongs to the selected course and semester
                    if str(student.course_id) != str(selected_course_id) or str(student.semester_id) != str(selected_semester_id):
                        return JsonResponse({
                            'status': 'session_mismatch',
                            'name': student.name,
                            'message': f'Choose correct session – student not in selected course/semester. Student is in {student.course.name} - {student.semester.name}.',
                            'student_course': student.course.name if student.course else 'N/A',
                            'student_semester': student.semester.name if student.semester else 'N/A'
                        })
                    
                    current_datetime = timezone.localtime()
                    today = current_datetime.date()
                    now_time = current_datetime.time()

                    # Try to find an existing attendance record for today
                    try:
                        record = AttendanceRecord.objects.get(
                            student=student,
                            date=today,
                            check_in_time__isnull=False  # Must have checked in
                        )
                        
                        if record.check_out_time:
                            return JsonResponse({
                                'status': 'already_checked_out',
                                'name': student.name,
                                'message': 'Already checked out today.'
                            })

                        # Record check-out time
                        record.check_out_time = now_time
                        record.save()

                        return JsonResponse({
                            'status': 'success',
                            'name': student.name,
                            'message': f'Successfully checked out at {now_time.strftime("%I:%M %p")}'
                        })

                    except AttendanceRecord.DoesNotExist:
                        return JsonResponse({
                            'status': 'error',
                            'message': 'No check-in record found for today.'
                        })

                except User.DoesNotExist:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Student not found in database.'
                    })
            else:
                print(f"[Checkout Debug] No match found. Best distance {best_distance:.4f} > tolerance 0.45")

        return JsonResponse({
            'status': 'not_recognized',
            'message': 'Face detected but not recognized.'
        })

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data received.'}, status=400)
    except Exception as e:
        import traceback
        print("------ UNEXPECTED ERROR in checkout_student_view ------")
        traceback.print_exc()
        print("-----------------------------------------------------")
        return JsonResponse({'status': 'error', 'message': 'An internal server error occurred.'}, status=500)

# --- VIEW FOR FACE RECOGNITION API ---
@csrf_exempt 
@require_POST 
@login_required 
@user_passes_test(is_admin) 
def recognize_face_view(request):
    """Receives image data, performs recognition, and marks attendance."""

    # Get selected course and semester from session
    selected_course_id = request.session.get('selected_course_id')
    selected_semester_id = request.session.get('selected_semester_id')
    
    if not selected_course_id or not selected_semester_id:
        return JsonResponse({
            'status': 'error', 
            'message': 'No course and semester selected for attendance session.'
        }, status=400)

    # --- Load known faces lazily or ensure they are loaded ---
    # Load if empty, or periodically refresh if needed
    if not known_face_encodings_global:
         print("[Recognition View] known_face_encodings_global is empty. Attempting to load...")
         load_known_faces()
         if not known_face_encodings_global:
              print("[Recognition View Error] Failed to load known faces.")
              return JsonResponse({'status': 'error', 'message': 'Known face encodings not loaded on server.'}, status=500)


    try:
        data = json.loads(request.body)
        image_data_uri = data.get('image_data')

        if not image_data_uri or ',' not in image_data_uri:
            return JsonResponse({'status': 'error', 'message': 'Invalid image data format.'}, status=400)

        # Decode base64 image data
        try:
            header, encoded = image_data_uri.split(',', 1)
            image_data = base64.b64decode(encoded)
        except (ValueError, TypeError) as e:
             return JsonResponse({'status': 'error', 'message': f'Error decoding base64 image: {e}'}, status=400)
        
        # Convert to OpenCV format (numpy array)
        nparr = np.frombuffer(image_data, np.uint8)
        frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        if frame is None:
             return JsonResponse({'status': 'error', 'message': 'Could not decode image data into frame.'}, status=400)

        # Convert image from BGR (OpenCV default) to RGB (face_recognition default)
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        
        # Find face locations and encodings in the current frame
        # Use "hog" for CPU (faster), "cnn" for GPU (more accurate)
        face_locations = face_recognition.face_locations(rgb_frame, model="hog") 
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)


        if not face_encodings:
             return JsonResponse({'status': 'no_face', 'message': 'No face detected.'})

        recognized_student_id = None
        recognized_student_name = "Unknown"
        
        # --- Face Matching Logic ---
        # Compare the first detected face against known faces
        TOLERANCE = 0.45  # Stricter tolerance for better accuracy
        matches = face_recognition.compare_faces(known_face_encodings_global, face_encodings[0], tolerance=TOLERANCE)
        face_distances = face_recognition.face_distance(known_face_encodings_global, face_encodings[0])
        
        if len(face_distances) > 0: 
            best_match_index = np.argmin(face_distances)
            best_distance = face_distances[best_match_index]
            
            # Debug logging for face matching
            print(f"[Face Match Debug] Best match index: {best_match_index}")
            print(f"[Face Match Debug] Best distance: {best_distance:.4f}")
            print(f"[Face Match Debug] Tolerance check: {best_distance <= TOLERANCE}")
            print(f"[Face Match Debug] All distances: {[f'{d:.4f}' for d in face_distances[:5]]}")  # Show first 5
            
            if matches[best_match_index]:
                recognized_student_id = known_face_ids_global[best_match_index]
                try:
                    student = User.objects.get(id=recognized_student_id)
                    recognized_student_name = student.name
                    print(f"[Face Match Debug] Matched student: {student.name} (ID: {student.id})")
                except User.DoesNotExist:
                     recognized_student_id = None
                     recognized_student_name = "ID Error"
                     print(f"[Face Match Error] Recognized student ID {recognized_student_id} not found.")
            else:
                print(f"[Face Match Debug] No match found. Best distance {best_distance:.4f} > tolerance {TOLERANCE}")
        else:
             print("[Recognition Warning] face_distances array was empty.")
        


        # --- Attendance Marking Logic ---
        if recognized_student_id:
            today = timezone.localdate()
            now_time = timezone.localtime().time()
            
            try:
                student = User.objects.get(id=recognized_student_id)
                
                # VALIDATE: Check if student belongs to the selected course and semester
                if str(student.course_id) != str(selected_course_id) or str(student.semester_id) != str(selected_semester_id):
                    return JsonResponse({
                        'status': 'session_mismatch',
                        'name': student.name,
                        'message': f'Choose correct session – student not in selected course/semester. Student is in {student.course.name} - {student.semester.name}.',
                        'student_course': student.course.name if student.course else 'N/A',
                        'student_semester': student.semester.name if student.semester else 'N/A'
                    })
                
                # Get current time in the correct timezone
                current_datetime = timezone.localtime()
                today = current_datetime.date()
                now_time = current_datetime.time()

                # Get cutoff times from settings
                time_settings = AttendanceSettings.get_instance()
                present_cutoff_time = time_settings.present_cutoff
                late_cutoff_time = time_settings.late_cutoff
                
                # Determine attendance status based on time
                if now_time <= present_cutoff_time:
                    status = 'present'
                elif now_time <= late_cutoff_time:
                    status = 'late'
                else:
                    status = 'absent'
                
                # Use get_or_create: marks attendance only on the *first* recognition of the day
                record, created = AttendanceRecord.objects.get_or_create(
                    student=student,
                    date=today,
                    defaults={
                        'status': status,
                        'check_in_time': now_time,
                        'manually_marked': False
                    }
                )

                # If record exists but no check-in time, update it
                if not created and not record.check_in_time:
                    record.check_in_time = now_time
                    record.status = status
                    record.save()

                if created:
                    attendance_status = f"Marked {status.title()}"
                else:
                    # If already marked, don't change the status
                    attendance_status = f"Already marked {record.status.title()}"

                return JsonResponse({
                    'status': 'success',
                    'name': recognized_student_name,
                    'user_id': recognized_student_id,
                    'attendance_status': attendance_status
                })
            except User.DoesNotExist:
                 return JsonResponse({'status': 'error', 'message': 'Recognized student ID not found.'}, status=500)
            except Exception as e:
                 print(f"[Attendance Marking Error] {e}")
                 return JsonResponse({'status': 'error', 'message': 'Error marking attendance.'}, status=500)
        else:
            # Face detected but not matched
            return JsonResponse({'status': 'not_recognized', 'message': 'Face detected, but not recognized.'})

    except json.JSONDecodeError:
        return JsonResponse({'status': 'error', 'message': 'Invalid JSON data received.'}, status=400)
    except Exception as e:
        import traceback
        print("------ UNEXPECTED ERROR in recognize_face_view ------")
        traceback.print_exc()
        print("-----------------------------------------------------")
        return JsonResponse({'status': 'error', 'message': f'An internal server error occurred.'}, status=500)

# -----------------------
# Student Views
# -----------------------

@login_required
@user_passes_test(is_student)
def leave_status_view(request):
    """Display student's leave request status."""
    student = request.user
    today = timezone.localdate()
    
    # Get all leave requests for the student
    all_leave_requests = LeaveRequest.objects.filter(student=student)
    
    # Get active leave (currently on leave)
    active_leave = all_leave_requests.filter(
        start_date__lte=today,
        end_date__gte=today,
        status='approved'
    ).first()

    # Get recent and pending leaves
    recent_leaves = all_leave_requests.order_by('-created_at')[:5]
    pending_leaves = all_leave_requests.filter(status='pending').order_by('-created_at')
    upcoming_leaves = all_leave_requests.filter(
        status='approved',
        start_date__gt=today
    ).order_by('start_date')[:3]

    # Calculate statistics
    leave_stats = {
        'total': all_leave_requests.count(),
        'pending': pending_leaves.count(),
        'approved': all_leave_requests.filter(status='approved').count(),
        'rejected': all_leave_requests.filter(status='rejected').count(),
    }

    context = {
        'active_leave': active_leave,
        'recent_leaves': recent_leaves,
        'pending_leaves': pending_leaves,
        'upcoming_leaves': upcoming_leaves,
        'leave_stats': leave_stats,
        'has_pending_leaves': pending_leaves.exists(),
        'has_active_leave': active_leave is not None,
    }
    
    return render(request, "core/leave_status.html", context)

@login_required
@user_passes_test(is_student)
def student_dashboard(request):
    """Displays the student's personalized dashboard."""
    from datetime import datetime, timedelta
    
    student = request.user
    today = timezone.localdate()
    
    # Fetch student's attendance records
    all_records = student.attendance_records.all()
    recent_attendance_records = all_records.order_by("-date")[:5]  # Show 5 most recent
    
    # Add duration calculation to each record
    recent_attendance = []
    for record in recent_attendance_records:
        record_data = {
            'date': record.date,
            'status': record.status,
            'check_in_time': record.check_in_time,
            'check_out_time': record.check_out_time,
            'duration': None
        }
        
        # Calculate duration if both check-in and check-out times exist
        if record.check_in_time and record.check_out_time:
            # Convert times to datetime for calculation
            check_in_dt = datetime.combine(record.date, record.check_in_time)
            check_out_dt = datetime.combine(record.date, record.check_out_time)
            
            # Calculate duration
            duration = check_out_dt - check_in_dt
            hours = int(duration.total_seconds() // 3600)
            minutes = int((duration.total_seconds() % 3600) // 60)
            
            if hours > 0:
                record_data['duration'] = f"{hours}h {minutes}m"
            else:
                record_data['duration'] = f"{minutes}m"
        
        recent_attendance.append(record_data)

    # Calculate attendance stats
    total_records = all_records.count()
    present_count = all_records.filter(status="present").count()
    absent_count = all_records.filter(status="absent").count()
    late_count = all_records.filter(status="late").count()
    percentage = round((present_count / total_records) * 100, 1) if total_records > 0 else 0

    context = {
        "student": student,
        "recent_attendance": recent_attendance,
        "attendance_stats": {
            "percentage": percentage,
            "present": present_count,
            "absent": absent_count,
            "late": late_count,
        }
    }
    return render(request, "core/student_dashboard.html", context)


@login_required
@user_passes_test(is_student)
def attendance_view(request):
    """Displays the student's full attendance history."""
    student = request.user
    # Order records by date, newest first
    records = student.attendance_records.order_by("-date")
    # TODO: Consider adding pagination if history can be long
    return render(request, "core/attendance.html", {"records": records})


@login_required
def courses_view(request):
    """Displays course information with different views for admin and students."""
    if request.user.is_admin:
        # Admin view - show all courses with student counts
        courses = Course.objects.select_related('department').all()
        course_stats = []
        
        for course in courses:
            students = User.objects.filter(course=course, is_student=True)
            active_students = students.filter(is_active=True)
            course_stats.append({
                'course': course,
                'total_students': students.count(),
                'active_students': active_students.count(),
                'inactive_students': students.filter(is_active=False).count(),
                'department': course.department,
                'recent_leaves': LeaveRequest.objects.filter(
                    student__course=course,
                    status='pending'
                ).count()
            })
        
        context = {
            'course_stats': course_stats,
            'total_courses': len(course_stats),
            'is_admin': True
        }
    else:
        # Student view - show their course and classmates
        student_course = request.user.course
        if student_course:
            classmates = User.objects.filter(
                course=student_course,
                is_student=True,
                is_active=True
            ).exclude(id=request.user.id)
            
            context = {
                'course': student_course,
                'department': student_course.department,
                'classmates': classmates,
                'classmate_count': classmates.count(),
                'is_admin': False
            }
        else:
            context = {'course': None, 'is_admin': False}
            messages.warning(request, "You are not enrolled in any course.")
    
    return render(request, "core/courses.html", context)


@login_required
@user_passes_test(is_student)
def apply_leave_view(request):
    """Handles the student's leave application form with comprehensive error handling."""
    from django.db import transaction
    print("\n=== Starting apply_leave_view ===")
    print(f"User: {request.user.username}, Method: {request.method}")  # Debug info
    context = {
        'form_data': request.POST if request.method == "POST" else {},
        'max_leave_days': 30  # Maximum days of leave allowed
    }
    
    if request.method == "POST":
        try:
            # Get and validate required fields
            date_from = request.POST.get("date_from", "").strip()
            date_to = request.POST.get("date_to", "").strip()
            reason = request.POST.get("reason", "").strip()

            # Check for empty fields with specific messages
            if not date_from:
                messages.error(request, "Please select a start date for your leave.")
                return render(request, "core/apply_leave.html", context)
            if not date_to:
                messages.error(request, "Please select an end date for your leave.")
                return render(request, "core/apply_leave.html", context)
            if not reason:
                messages.error(request, "Please provide a reason for your leave request.")
                return render(request, "core/apply_leave.html", context)

            # Validate date formats
            try:
                start_date = datetime.strptime(date_from, "%Y-%m-%d").date()
                end_date = datetime.strptime(date_to, "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "Please enter valid dates in DD/MM/YYYY format.")
                return render(request, "core/apply_leave.html", context)

            today = timezone.localdate()

            # Comprehensive date validations
            if start_date < today:
                messages.error(request, "Leave cannot be requested for past dates. Please select a future date.")
                return render(request, "core/apply_leave.html", context)

            if end_date < start_date:
                messages.error(request, "End date must be after or equal to the start date.")
                return render(request, "core/apply_leave.html", context)

            leave_duration = (end_date - start_date).days + 1
            if leave_duration > context['max_leave_days']:
                messages.error(request, f"Leave request exceeds maximum allowed duration of {context['max_leave_days']} days.")
                return render(request, "core/apply_leave.html", context)

            # Check for leave request limit
            month_start = today.replace(day=1)
            month_end = (today.replace(day=28) + timezone.timedelta(days=4)).replace(day=1) - timezone.timedelta(days=1)
            month_requests = LeaveRequest.objects.filter(
                student=request.user,
                created_at__date__range=[month_start, month_end]
            ).count()

            if month_requests >= 3:
                messages.error(request, "You have reached the maximum number of leave requests (3) for this month.")
                return render(request, "core/apply_leave.html", context)

            # Check for overlapping leaves
            overlapping_leaves = LeaveRequest.objects.filter(
                student=request.user,
                status__in=['pending', 'approved'],
                start_date__lte=end_date,
                end_date__gte=start_date
            )

            if overlapping_leaves.exists():
                overlap = overlapping_leaves.first()
                messages.error(request, 
                    f"You already have a {overlap.status} leave request from "
                    f"{overlap.start_date.strftime('%d %b, %Y')} to {overlap.end_date.strftime('%d %b, %Y')}.")
                return render(request, "core/apply_leave.html", context)

            # Check reason length and word count
            word_count = len(reason.split())
            if len(reason) < 10 or word_count < 10:
                messages.error(request, 
                    "Please provide a more detailed reason for your leave request. "
                    "Your reason should be at least 10 words long. "
                    f"Current length: {word_count} words.")
                context['form_data'] = request.POST  # Preserve form data
                context['error_field'] = 'reason'  # Highlight the reason field
                return render(request, "core/apply_leave.html", context)

            # All validations passed, create the leave request
            try:
                print("\n=== Creating Leave Request ===")
                print(f"Student: {request.user.username}")
                print(f"Start Date: {start_date}")
                print(f"End Date: {end_date}")
                print(f"Reason: {reason}")

                # First, verify the student user
                if not hasattr(request.user, 'is_student') or not request.user.is_student:
                    raise ValueError("Current user is not a student")

                # Create the leave request with transaction
                with transaction.atomic():
                    leave_request = LeaveRequest.objects.create(
                        student=request.user,
                        start_date=start_date,
                        end_date=end_date,
                        reason=reason,
                        status='pending',
                        created_at=timezone.now()
                    )
                    
                    # Force a save
                    leave_request.save()
                    
                    print(f"\nLeave Request Created:")
                    print(f"ID: {leave_request.id}")
                    print(f"Student: {leave_request.student.username}")
                    print(f"Status: {leave_request.status}")
                    
                    # Verify in database
                    verify = LeaveRequest.objects.filter(id=leave_request.id).exists()
                    print(f"\nVerification:")
                    print(f"Request exists in database: {verify}")
                    
                    if verify:
                        # Count total leave requests for this student
                        student_requests = LeaveRequest.objects.filter(student=request.user).count()
                        print(f"Total leave requests for student: {student_requests}")
                        
                        # Success message
                        messages.success(request, 
                            f"Leave request submitted successfully for {leave_duration} day(s) "
                            f"from {start_date.strftime('%d %b, %Y')} to {end_date.strftime('%d %b, %Y')}. "
                            "You will be notified when it is reviewed.")
                    
                # Immediately notify admins of new leave request
                try:
                    admin_users = User.objects.filter(is_admin=True)
                    for admin in admin_users:
                        messages.info(admin, f"New leave request from {request.user.name} is pending review.")
                except Exception as e:
                    print(f"Error notifying admins: {e}")
                    
            except Exception as e:
                print(f"Error creating leave request: {e}")
                messages.error(request, "Error submitting leave request. Please try again.")
            
            # Try to notify admins via email
            try:
                admin_users = User.objects.filter(is_admin=True)
                admin_emails = [user.email for user in admin_users if user.email]
                
                if admin_emails:
                    subject = f"New Leave Request: {request.user.name}"
                    message = f"""A new leave request has been submitted:
                    
Student: {request.user.name}
Roll Number: {request.user.roll_no}
Duration: {start_date.strftime('%d %b, %Y')} to {end_date.strftime('%d %b, %Y')}
Reason: {reason}

Please review this request at:
{request.build_absolute_uri(reverse('manage_leave'))}"""
                    
                    # Send email notification
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        admin_emails,
                        fail_silently=True
                    )
                    
                    print(f"Notification sent to admin emails: {admin_emails}")
                    
                    # Also update admin's session data to show pending requests
                    for admin in admin_users:
                        try:
                            if 'pending_leave_requests' in request.session:
                                request.session['pending_leave_requests'] += 1
                            else:
                                request.session['pending_leave_requests'] = 1
                            request.session.modified = True
                        except Exception as session_error:
                            print(f"Error updating admin session: {session_error}")
                            
            except Exception as e:
                print(f"Failed to send admin notification email: {e}")

            return redirect('student_dashboard')

        except Exception as e:
            print(f"Unexpected error in apply_leave_view: {e}")
            messages.error(request, "An unexpected error occurred. Please try again or contact support if the problem persists.")
            return render(request, "core/apply_leave.html", context)

    # For GET request, show existing leave requests and the form
    try:
        # Get all leave requests for the current student
        user_leaves = LeaveRequest.objects.filter(
            student=request.user
        ).order_by('-created_at')

        context.update({
            'leave_requests': user_leaves,
            'pending_count': user_leaves.filter(status='pending').count(),
            'approved_count': user_leaves.filter(status='approved').count(),
            'rejected_count': user_leaves.filter(status='rejected').count()
        })
        
        return render(request, "core/apply_leave.html", context)
    
    except Exception as e:
        print(f"Error fetching leave requests: {e}")
        messages.error(request, "Unable to load your leave requests. Please refresh the page or contact support.")
        return render(request, "core/apply_leave.html", context)

@login_required
@user_passes_test(is_admin)
def download_attendance_report(request):
    """Download attendance records as Excel file based on current filters."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    # Get the same filter parameters as attendance_details_view
    student_search = request.GET.get('student_search', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    status = request.GET.get('status', '')
    department_id = request.GET.get('department', '')
    course_id = request.GET.get('course', '')
    session_id = request.GET.get('session', '')
    semester_id = request.GET.get('semester', '')

    # Check if required filters are applied
    has_required_filters = department_id and course_id and semester_id
    
    if not has_required_filters:
        messages.error(request, "Please select Department, Course, and Semester to download attendance report.")
        return redirect('attendance_details')
    
    # Get filtered records (same logic as attendance_details_view)
    records = AttendanceRecord.objects.all().select_related('student__department', 'student__course', 'student__session', 'student__semester')
    
    # Apply mandatory filters
    records = records.filter(
        student__department_id=department_id,
        student__course_id=course_id,
        student__semester_id=semester_id
    )
    
    # Apply additional filters
    if student_search:
        records = records.filter(
            Q(student__name__icontains=student_search) |
            Q(student__roll_no__icontains=student_search)
        )

    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            records = records.filter(date__gte=start_date)
        except ValueError:
            pass

    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            records = records.filter(date__lte=end_date)
        except ValueError:
            pass

    if status and status in dict(AttendanceRecord.STATUS_CHOICES):
        records = records.filter(status=status)

    if session_id:
        records = records.filter(student__session_id=session_id)

    # Order by date and student name
    records = records.order_by('-date', 'student__name')
    
    # Get filter names for the report
    try:
        department = Department.objects.get(id=department_id)
        course = Course.objects.get(id=course_id)
        semester = Semester.objects.get(id=semester_id)
        session = None
        if session_id:
            session = Session.objects.get(id=session_id)
    except:
        messages.error(request, "Invalid filter parameters.")
        return redirect('attendance_details')
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Report title
    ws['A1'] = "ATTENDANCE REPORT"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:I1')
    
    # Report info
    current_row = 3
    ws[f'A{current_row}'] = "Department:"
    ws[f'B{current_row}'] = department.name
    ws[f'A{current_row}'].font = Font(bold=True)
    
    ws[f'D{current_row}'] = "Course:"
    ws[f'E{current_row}'] = course.name
    ws[f'D{current_row}'].font = Font(bold=True)
    
    ws[f'G{current_row}'] = "Semester:"
    ws[f'H{current_row}'] = semester.name
    ws[f'G{current_row}'].font = Font(bold=True)
    
    current_row += 1
    if session:
        ws[f'A{current_row}'] = "Session:"
        ws[f'B{current_row}'] = session.year
        ws[f'A{current_row}'].font = Font(bold=True)
    
    if start_date or end_date:
        date_range = ""
        if start_date and end_date:
            date_range = f"{start_date} to {end_date}"
        elif start_date:
            date_range = f"From {start_date}"
        elif end_date:
            date_range = f"Until {end_date}"
        
        ws[f'D{current_row}'] = "Date Range:"
        ws[f'E{current_row}'] = date_range
        ws[f'D{current_row}'].font = Font(bold=True)
    
    ws[f'G{current_row}'] = "Generated:"
    ws[f'H{current_row}'] = timezone.now().strftime("%Y-%m-%d %H:%M:%S")
    ws[f'G{current_row}'].font = Font(bold=True)
    
    current_row += 2
    
    # Column headers
    headers = [
        'Student Name', 'Roll Number', 'Department', 'Course', 'Session', 'Semester', 
        'Date', 'Status', 'Check-In Time', 'Check-Out Time', 'Method'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    current_row += 1
    
    # Data rows
    for record in records:
        row_data = [
            record.student.name,
            record.student.roll_no,
            record.student.department.name,
            record.student.course.name,
            record.student.session.year if record.student.session else '',
            record.student.semester.name,
            record.date.strftime('%Y-%m-%d'),
            record.status.title(),
            record.check_in_time.strftime('%I:%M %p') if record.check_in_time else '',
            record.check_out_time.strftime('%I:%M %p') if record.check_out_time else '',
            'Manual' if record.manually_marked else 'System (Face)'
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = value
            cell.border = border
            if col_num in [8, 9, 10]:  # Center align time columns
                cell.alignment = center_align
        
        current_row += 1
    
    # Summary section
    current_row += 2
    ws[f'A{current_row}'] = "SUMMARY"
    ws[f'A{current_row}'].font = Font(bold=True, size=14)
    ws.merge_cells(f'A{current_row}:C{current_row}')
    
    current_row += 1
    total_records = records.count()
    present_count = records.filter(status='present').count()
    absent_count = records.filter(status='absent').count()
    late_count = records.filter(status='late').count()
    
    summary_data = [
        ('Total Records', total_records),
        ('Present', present_count),
        ('Absent', absent_count),
        ('Late', late_count),
        ('Attendance Rate', f"{(present_count / total_records * 100):.1f}%" if total_records > 0 else "0%")
    ]
    
    for label, value in summary_data:
        ws[f'A{current_row}'] = label + ":"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = value
        current_row += 1
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row_num in range(1, current_row):
            cell_value = str(ws[f'{column_letter}{row_num}'].value or '')
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename
    filename_parts = [department.name, course.name, semester.name]
    if session:
        filename_parts.append(str(session.year))
    filename = f"attendance_report_{'_'.join(filename_parts)}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response
